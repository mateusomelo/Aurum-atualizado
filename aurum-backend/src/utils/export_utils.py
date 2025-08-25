from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import xlsxwriter
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
import io
from datetime import datetime

class ReportExporter:
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        self.subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_LEFT,
            textColor=colors.darkgreen
        )
        
    def export_empresas_pdf(self, relatorio_data):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Título
        title = Paragraph("Relatório de Empresas", self.title_style)
        story.append(title)
        
        # Data de geração
        date_str = datetime.now().strftime('%d/%m/%Y às %H:%M')
        date_p = Paragraph(f"<i>Gerado em: {date_str}</i>", self.styles['Normal'])
        story.append(date_p)
        story.append(Spacer(1, 20))
        
        # Resumo geral
        total_empresas = len(relatorio_data)
        total_usuarios = sum(item['total_usuarios'] for item in relatorio_data)
        total_chamados = sum(item['total_chamados'] for item in relatorio_data)
        total_finalizados = sum(item['chamados_finalizados'] for item in relatorio_data)
        
        resumo_data = [
            ['Métrica', 'Valor'],
            ['Total de Empresas Ativas', str(total_empresas)],
            ['Total de Usuários', str(total_usuarios)],
            ['Total de Chamados', str(total_chamados)],
            ['Chamados Finalizados', str(total_finalizados)],
            ['Taxa de Resolução Geral', f"{(total_finalizados/total_chamados*100) if total_chamados > 0 else 0:.1f}%"]
        ]
        
        resumo_table = Table(resumo_data)
        resumo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(Paragraph("Resumo Geral", self.subtitle_style))
        story.append(resumo_table)
        story.append(Spacer(1, 30))
        
        # Detalhes por empresa
        story.append(Paragraph("Detalhes por Empresa", self.subtitle_style))
        
        for empresa_data in relatorio_data:
            # Nome da empresa
            empresa_title = Paragraph(f"<b>{empresa_data['empresa'].nome_empresa}</b>", 
                                    ParagraphStyle('EmpresaTitle', fontSize=12, textColor=colors.darkblue))
            story.append(empresa_title)
            
            # Info básica da empresa
            info = f"CNPJ: {empresa_data['empresa'].cnpj} | Organizador: {empresa_data['empresa'].organizador}"
            story.append(Paragraph(info, self.styles['Normal']))
            story.append(Spacer(1, 10))
            
            # Estatísticas da empresa
            stats_data = [
                ['Métrica', 'Quantidade'],
                ['Total de Usuários', str(empresa_data['total_usuarios'])],
                ['Total de Chamados', str(empresa_data['total_chamados'])],
                ['Chamados Abertos', str(empresa_data['chamados_abertos'])],
                ['Chamados em Andamento', str(empresa_data['chamados_andamento'])],
                ['Chamados Finalizados', str(empresa_data['chamados_finalizados'])],
                ['Taxa de Resolução', f"{(empresa_data['chamados_finalizados']/empresa_data['total_chamados']*100) if empresa_data['total_chamados'] > 0 else 0:.1f}%"]
            ]
            
            stats_table = Table(stats_data)
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(stats_table)
            
            # Usuários que abriram chamados
            if empresa_data['usuarios_stats']:
                story.append(Spacer(1, 15))
                story.append(Paragraph("<b>Usuários que Abriram Chamados:</b>", self.styles['Normal']))
                
                users_data = [['Usuário', 'Email', 'Total', 'Abertos', 'Andamento', 'Finalizados']]
                for user_stat in empresa_data['usuarios_stats']:
                    if user_stat['total_chamados'] > 0:
                        users_data.append([
                            user_stat['usuario'].nome,
                            user_stat['usuario'].email,
                            str(user_stat['total_chamados']),
                            str(user_stat['abertos']),
                            str(user_stat['em_andamento']),
                            str(user_stat['finalizados'])
                        ])
                
                if len(users_data) > 1:
                    users_table = Table(users_data)
                    users_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.aliceblue),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(users_table)
            
            story.append(Spacer(1, 30))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def export_empresas_excel(self, relatorio_data):
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        
        # Formatos
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#1f4e79',
            'font_color': 'white'
        })
        
        header_format = workbook.add_format({
            'bold': True,
            'font_size': 12,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#2e75b6',
            'font_color': 'white',
            'border': 1
        })
        
        data_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        # Aba Resumo
        resumo_sheet = workbook.add_worksheet('Resumo Geral')
        
        resumo_sheet.write('A1', 'RELATÓRIO DE EMPRESAS', title_format)
        resumo_sheet.merge_range('A1:F1', 'RELATÓRIO DE EMPRESAS', title_format)
        
        # Data de geração
        date_str = datetime.now().strftime('%d/%m/%Y às %H:%M')
        resumo_sheet.write('A2', f'Gerado em: {date_str}')
        
        # Estatísticas gerais
        resumo_sheet.write('A4', 'Métrica', header_format)
        resumo_sheet.write('B4', 'Valor', header_format)
        
        total_empresas = len(relatorio_data)
        total_usuarios = sum(item['total_usuarios'] for item in relatorio_data)
        total_chamados = sum(item['total_chamados'] for item in relatorio_data)
        total_finalizados = sum(item['chamados_finalizados'] for item in relatorio_data)
        
        metrics = [
            ['Total de Empresas Ativas', total_empresas],
            ['Total de Usuários', total_usuarios],
            ['Total de Chamados', total_chamados],
            ['Chamados Finalizados', total_finalizados],
            ['Taxa de Resolução Geral', f"{(total_finalizados/total_chamados*100) if total_chamados > 0 else 0:.1f}%"]
        ]
        
        for i, (metric, value) in enumerate(metrics, 5):
            resumo_sheet.write(f'A{i}', metric, data_format)
            resumo_sheet.write(f'B{i}', value, data_format)
        
        # Ajustar largura das colunas
        resumo_sheet.set_column('A:A', 25)
        resumo_sheet.set_column('B:B', 20)
        
        # Aba por empresa
        empresas_sheet = workbook.add_worksheet('Empresas Detalhado')
        
        # Cabeçalhos
        headers = ['Empresa', 'CNPJ', 'Organizador', 'Total Usuários', 'Total Chamados', 
                  'Abertos', 'Em Andamento', 'Finalizados', 'Taxa Resolução (%)']
        
        for i, header in enumerate(headers):
            empresas_sheet.write(0, i, header, header_format)
        
        # Dados das empresas
        for row, empresa_data in enumerate(relatorio_data, 1):
            empresas_sheet.write(row, 0, empresa_data['empresa'].nome_empresa, data_format)
            empresas_sheet.write(row, 1, empresa_data['empresa'].cnpj, data_format)
            empresas_sheet.write(row, 2, empresa_data['empresa'].organizador, data_format)
            empresas_sheet.write(row, 3, empresa_data['total_usuarios'], data_format)
            empresas_sheet.write(row, 4, empresa_data['total_chamados'], data_format)
            empresas_sheet.write(row, 5, empresa_data['chamados_abertos'], data_format)
            empresas_sheet.write(row, 6, empresa_data['chamados_andamento'], data_format)
            empresas_sheet.write(row, 7, empresa_data['chamados_finalizados'], data_format)
            
            taxa = (empresa_data['chamados_finalizados']/empresa_data['total_chamados']*100) if empresa_data['total_chamados'] > 0 else 0
            empresas_sheet.write(row, 8, f"{taxa:.1f}%", data_format)
        
        # Ajustar largura das colunas
        empresas_sheet.set_column('A:A', 25)
        empresas_sheet.set_column('B:B', 15)
        empresas_sheet.set_column('C:C', 20)
        empresas_sheet.set_column('D:I', 12)
        
        workbook.close()
        buffer.seek(0)
        return buffer
    
    def export_tecnicos_pdf(self, relatorio_data):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Título
        title = Paragraph("Relatório de Técnicos", self.title_style)
        story.append(title)
        
        # Data de geração
        date_str = datetime.now().strftime('%d/%m/%Y às %H:%M')
        date_p = Paragraph(f"<i>Gerado em: {date_str}</i>", self.styles['Normal'])
        story.append(date_p)
        story.append(Spacer(1, 20))
        
        # Resumo geral
        total_tecnicos = len(relatorio_data)
        total_chamados = sum(item['total_chamados'] for item in relatorio_data)
        total_andamento = sum(item['chamados_andamento'] for item in relatorio_data)
        total_finalizados = sum(item['chamados_finalizados'] for item in relatorio_data)
        
        resumo_data = [
            ['Métrica', 'Valor'],
            ['Total de Técnicos Ativos', str(total_tecnicos)],
            ['Total de Chamados Atendidos', str(total_chamados)],
            ['Chamados em Andamento', str(total_andamento)],
            ['Chamados Finalizados', str(total_finalizados)],
            ['Taxa de Resolução Geral', f"{(total_finalizados/total_chamados*100) if total_chamados > 0 else 0:.1f}%"]
        ]
        
        resumo_table = Table(resumo_data)
        resumo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(Paragraph("Resumo Geral", self.subtitle_style))
        story.append(resumo_table)
        story.append(Spacer(1, 30))
        
        # Detalhes por técnico
        story.append(Paragraph("Detalhes por Técnico", self.subtitle_style))
        
        for tecnico_data in relatorio_data:
            # Nome do técnico
            tecnico_title = Paragraph(f"<b>{tecnico_data['tecnico'].nome}</b>", 
                                    ParagraphStyle('TecnicoTitle', fontSize=12, textColor=colors.darkgreen))
            story.append(tecnico_title)
            
            # Info básica do técnico
            info = f"Email: {tecnico_data['tecnico'].email} | Telefone: {tecnico_data['tecnico'].telefone}"
            story.append(Paragraph(info, self.styles['Normal']))
            story.append(Spacer(1, 10))
            
            # Estatísticas do técnico
            stats_data = [
                ['Métrica', 'Quantidade'],
                ['Total Atendidos', str(tecnico_data['total_chamados'])],
                ['Chamados Abertos', str(tecnico_data['chamados_abertos'])],
                ['Chamados em Andamento', str(tecnico_data['chamados_andamento'])],
                ['Chamados Finalizados', str(tecnico_data['chamados_finalizados'])],
                ['Taxa de Resolução', f"{(tecnico_data['chamados_finalizados']/tecnico_data['total_chamados']*100) if tecnico_data['total_chamados'] > 0 else 0:.1f}%"]
            ]
            
            stats_table = Table(stats_data)
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(stats_table)
            story.append(Spacer(1, 30))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def export_tecnicos_excel(self, relatorio_data):
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        
        # Formatos
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#2e7d2e',
            'font_color': 'white'
        })
        
        header_format = workbook.add_format({
            'bold': True,
            'font_size': 12,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#5cb85c',
            'font_color': 'white',
            'border': 1
        })
        
        data_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        # Aba Resumo
        resumo_sheet = workbook.add_worksheet('Resumo Geral')
        
        resumo_sheet.write('A1', 'RELATÓRIO DE TÉCNICOS', title_format)
        resumo_sheet.merge_range('A1:F1', 'RELATÓRIO DE TÉCNICOS', title_format)
        
        # Data de geração
        date_str = datetime.now().strftime('%d/%m/%Y às %H:%M')
        resumo_sheet.write('A2', f'Gerado em: {date_str}')
        
        # Estatísticas gerais
        resumo_sheet.write('A4', 'Métrica', header_format)
        resumo_sheet.write('B4', 'Valor', header_format)
        
        total_tecnicos = len(relatorio_data)
        total_chamados = sum(item['total_chamados'] for item in relatorio_data)
        total_andamento = sum(item['chamados_andamento'] for item in relatorio_data)
        total_finalizados = sum(item['chamados_finalizados'] for item in relatorio_data)
        
        metrics = [
            ['Total de Técnicos Ativos', total_tecnicos],
            ['Total de Chamados Atendidos', total_chamados],
            ['Chamados em Andamento', total_andamento],
            ['Chamados Finalizados', total_finalizados],
            ['Taxa de Resolução Geral', f"{(total_finalizados/total_chamados*100) if total_chamados > 0 else 0:.1f}%"]
        ]
        
        for i, (metric, value) in enumerate(metrics, 5):
            resumo_sheet.write(f'A{i}', metric, data_format)
            resumo_sheet.write(f'B{i}', value, data_format)
        
        # Aba por técnico
        tecnicos_sheet = workbook.add_worksheet('Técnicos Detalhado')
        
        # Cabeçalhos
        headers = ['Técnico', 'Email', 'Telefone', 'Total Atendidos', 
                  'Abertos', 'Em Andamento', 'Finalizados', 'Taxa Resolução (%)']
        
        for i, header in enumerate(headers):
            tecnicos_sheet.write(0, i, header, header_format)
        
        # Dados dos técnicos
        for row, tecnico_data in enumerate(relatorio_data, 1):
            tecnicos_sheet.write(row, 0, tecnico_data['tecnico'].nome, data_format)
            tecnicos_sheet.write(row, 1, tecnico_data['tecnico'].email, data_format)
            tecnicos_sheet.write(row, 2, tecnico_data['tecnico'].telefone, data_format)
            tecnicos_sheet.write(row, 3, tecnico_data['total_chamados'], data_format)
            tecnicos_sheet.write(row, 4, tecnico_data['chamados_abertos'], data_format)
            tecnicos_sheet.write(row, 5, tecnico_data['chamados_andamento'], data_format)
            tecnicos_sheet.write(row, 6, tecnico_data['chamados_finalizados'], data_format)
            
            taxa = (tecnico_data['chamados_finalizados']/tecnico_data['total_chamados']*100) if tecnico_data['total_chamados'] > 0 else 0
            tecnicos_sheet.write(row, 7, f"{taxa:.1f}%", data_format)
        
        # Ajustar largura das colunas
        tecnicos_sheet.set_column('A:A', 25)
        tecnicos_sheet.set_column('B:B', 25)
        tecnicos_sheet.set_column('C:C', 15)
        tecnicos_sheet.set_column('D:H', 12)
        
        workbook.close()
        buffer.seek(0)
        return buffer