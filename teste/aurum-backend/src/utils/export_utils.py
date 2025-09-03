from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import xlsxwriter
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
import io
from datetime import datetime

class ReportExporter:
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        self.subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_LEFT,
            textColor=colors.darkgreen
        )
        
    def export_empresas_pdf(self, relatorio_data):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Título
        title = Paragraph("Relatório de Empresas", self.title_style)
        story.append(title)
        
        # Data de geração
        date_str = datetime.now().strftime('%d/%m/%Y às %H:%M')
        date_p = Paragraph(f"<i>Gerado em: {date_str}</i>", self.styles['Normal'])
        story.append(date_p)
        story.append(Spacer(1, 20))
        
        # Resumo geral
        total_empresas = len(relatorio_data)
        total_usuarios = sum(item['total_usuarios'] for item in relatorio_data)
        total_chamados = sum(item['total_chamados'] for item in relatorio_data)
        total_finalizados = sum(item['chamados_finalizados'] for item in relatorio_data)
        
        resumo_data = [
            ['Métrica', 'Valor'],
            ['Total de Empresas Ativas', str(total_empresas)],
            ['Total de Usuários', str(total_usuarios)],
            ['Total de Chamados', str(total_chamados)],
            ['Chamados Finalizados', str(total_finalizados)],
            ['Taxa de Resolução Geral', f"{(total_finalizados/total_chamados*100) if total_chamados > 0 else 0:.1f}%"]
        ]
        
        resumo_table = Table(resumo_data)
        resumo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(Paragraph("Resumo Geral", self.subtitle_style))
        story.append(resumo_table)
        story.append(Spacer(1, 30))
        
        # Detalhes por empresa
        story.append(Paragraph("Detalhes por Empresa", self.subtitle_style))
        
        for empresa_data in relatorio_data:
            # Nome da empresa
            empresa_title = Paragraph(f"<b>{empresa_data['empresa'].nome_empresa}</b>", 
                                    ParagraphStyle('EmpresaTitle', fontSize=12, textColor=colors.darkblue))
            story.append(empresa_title)
            
            # Info básica da empresa
            info_lines = [
                f"CNPJ: {empresa_data['empresa'].cnpj}",
                f"Organizador: {empresa_data['empresa'].organizador}",
                f"Data de Cadastro: {empresa_data['empresa'].data_criacao.strftime('%d/%m/%Y') if hasattr(empresa_data['empresa'], 'data_criacao') else 'N/A'}",
                f"Telefone: {empresa_data['empresa'].telefone if hasattr(empresa_data['empresa'], 'telefone') and empresa_data['empresa'].telefone else 'N/A'}",
                f"Endereço: {empresa_data['empresa'].endereco if hasattr(empresa_data['empresa'], 'endereco') and empresa_data['empresa'].endereco else 'N/A'}"
            ]
            for info_line in info_lines:
                story.append(Paragraph(info_line, self.styles['Normal']))
            story.append(Spacer(1, 10))
            
            # Estatísticas da empresa
            stats_data = [
                ['Métrica', 'Quantidade'],
                ['Total de Usuários', str(empresa_data['total_usuarios'])],
                ['Total de Chamados', str(empresa_data['total_chamados'])],
                ['Chamados Abertos', str(empresa_data['chamados_abertos'])],
                ['Chamados em Andamento', str(empresa_data['chamados_andamento'])],
                ['Chamados Finalizados', str(empresa_data['chamados_finalizados'])],
                ['Taxa de Resolução', f"{(empresa_data['chamados_finalizados']/empresa_data['total_chamados']*100) if empresa_data['total_chamados'] > 0 else 0:.1f}%"]
            ]
            
            stats_table = Table(stats_data)
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(stats_table)
            
            # Usuários que abriram chamados
            if empresa_data['usuarios_stats']:
                story.append(Spacer(1, 15))
                story.append(Paragraph("<b>Usuários que Abriram Chamados:</b>", self.styles['Normal']))
                
                users_data = [['Usuário', 'Email', 'Total', 'Abertos', 'Andamento', 'Finalizados']]
                for user_stat in empresa_data['usuarios_stats']:
                    if user_stat['total_chamados'] > 0:
                        users_data.append([
                            user_stat['usuario'].nome,
                            user_stat['usuario'].email,
                            str(user_stat['total_chamados']),
                            str(user_stat['abertos']),
                            str(user_stat['em_andamento']),
                            str(user_stat['finalizados'])
                        ])
                
                if len(users_data) > 1:
                    users_table = Table(users_data)
                    users_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.aliceblue),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(users_table)
            
            # Técnicos que atenderam esta empresa
            if 'tecnicos_atenderam' in empresa_data and empresa_data['tecnicos_atenderam']:
                story.append(Spacer(1, 15))
                story.append(Paragraph("<b>Técnicos que Atenderam:</b>", self.styles['Normal']))
                
                tecnicos_data = [['Técnico', 'Email', 'Chamados Atendidos', 'Finalizados', 'Taxa (%)']]
                for tecnico_stat in empresa_data['tecnicos_atenderam']:
                    taxa = (tecnico_stat['finalizados']/tecnico_stat['chamados_atendidos']*100) if tecnico_stat['chamados_atendidos'] > 0 else 0
                    tecnicos_data.append([
                        tecnico_stat['tecnico'].nome,
                        tecnico_stat['tecnico'].email,
                        str(tecnico_stat['chamados_atendidos']),
                        str(tecnico_stat['finalizados']),
                        f"{taxa:.1f}%"
                    ])
                
                tecnicos_table = Table(tecnicos_data)
                tecnicos_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.purple),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lavender),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(tecnicos_table)
            
            # Lista detalhada de chamados (últimos 10)
            if 'chamados_detalhados' in empresa_data and empresa_data['chamados_detalhados']:
                story.append(Spacer(1, 15))
                story.append(Paragraph("<b>Últimos 10 Chamados Detalhados:</b>", self.styles['Normal']))
                
                chamados_data = [['Título', 'Status', 'Prioridade', 'Data Abertura', 'Data Finalização', 'Usuário', 'Técnico', 'Tempo (dias)']]
                # Mostrar apenas os últimos 10 chamados, ordenados por data
                ultimos_chamados = sorted(empresa_data['chamados_detalhados'], key=lambda x: x['data_abertura'], reverse=True)[:10]
                
                for chamado_detail in ultimos_chamados:
                    data_abertura = chamado_detail['data_abertura'].strftime('%d/%m/%Y %H:%M') if chamado_detail['data_abertura'] else 'N/A'
                    data_finalizacao = chamado_detail['data_finalizacao'].strftime('%d/%m/%Y %H:%M') if chamado_detail['data_finalizacao'] else 'Em aberto'
                    tempo_resolucao = f"{chamado_detail['tempo_resolucao']}" if chamado_detail['tempo_resolucao'] is not None else 'N/A'
                    
                    chamados_data.append([
                        chamado_detail['titulo'][:30] + '...' if len(chamado_detail['titulo']) > 30 else chamado_detail['titulo'],
                        chamado_detail['status'].replace('_', ' ').title(),
                        chamado_detail['prioridade'].title(),
                        data_abertura,
                        data_finalizacao,
                        chamado_detail['usuario'][:20] if len(chamado_detail['usuario']) > 20 else chamado_detail['usuario'],
                        chamado_detail['tecnico'][:15] if len(chamado_detail['tecnico']) > 15 else chamado_detail['tecnico'],
                        tempo_resolucao
                    ])
                
                if len(chamados_data) > 1:
                    chamados_table = Table(chamados_data)
                    chamados_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.darkred),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 6),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.mistyrose),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(chamados_table)
            
            story.append(Spacer(1, 30))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def export_empresas_excel(self, relatorio_data):
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        
        # Formatos
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#1f4e79',
            'font_color': 'white'
        })
        
        header_format = workbook.add_format({
            'bold': True,
            'font_size': 12,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#2e75b6',
            'font_color': 'white',
            'border': 1
        })
        
        data_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        # Aba Resumo
        resumo_sheet = workbook.add_worksheet('Resumo Geral')
        
        resumo_sheet.write('A1', 'RELATÓRIO DE EMPRESAS', title_format)
        resumo_sheet.merge_range('A1:F1', 'RELATÓRIO DE EMPRESAS', title_format)
        
        # Data de geração
        date_str = datetime.now().strftime('%d/%m/%Y às %H:%M')
        resumo_sheet.write('A2', f'Gerado em: {date_str}')
        
        # Estatísticas gerais
        resumo_sheet.write('A4', 'Métrica', header_format)
        resumo_sheet.write('B4', 'Valor', header_format)
        
        total_empresas = len(relatorio_data)
        total_usuarios = sum(item['total_usuarios'] for item in relatorio_data)
        total_chamados = sum(item['total_chamados'] for item in relatorio_data)
        total_finalizados = sum(item['chamados_finalizados'] for item in relatorio_data)
        
        metrics = [
            ['Total de Empresas Ativas', total_empresas],
            ['Total de Usuários', total_usuarios],
            ['Total de Chamados', total_chamados],
            ['Chamados Finalizados', total_finalizados],
            ['Taxa de Resolução Geral', f"{(total_finalizados/total_chamados*100) if total_chamados > 0 else 0:.1f}%"]
        ]
        
        for i, (metric, value) in enumerate(metrics, 5):
            resumo_sheet.write(f'A{i}', metric, data_format)
            resumo_sheet.write(f'B{i}', value, data_format)
        
        # Ajustar largura das colunas
        resumo_sheet.set_column('A:A', 25)
        resumo_sheet.set_column('B:B', 20)
        
        # Aba por empresa
        empresas_sheet = workbook.add_worksheet('Empresas Detalhado')
        
        # Cabeçalhos
        headers = ['Empresa', 'CNPJ', 'Organizador', 'Data Cadastro', 'Telefone', 'Total Usuários', 'Total Chamados', 
                  'Abertos', 'Em Andamento', 'Finalizados', 'Taxa Resolução (%)', 'Técnicos Atenderam']
        
        for i, header in enumerate(headers):
            empresas_sheet.write(0, i, header, header_format)
        
        # Dados das empresas
        for row, empresa_data in enumerate(relatorio_data, 1):
            empresas_sheet.write(row, 0, empresa_data['empresa'].nome_empresa, data_format)
            empresas_sheet.write(row, 1, empresa_data['empresa'].cnpj, data_format)
            empresas_sheet.write(row, 2, empresa_data['empresa'].organizador, data_format)
            empresas_sheet.write(row, 3, empresa_data['empresa'].data_criacao.strftime('%d/%m/%Y') if hasattr(empresa_data['empresa'], 'data_criacao') else 'N/A', data_format)
            empresas_sheet.write(row, 4, empresa_data['empresa'].telefone if hasattr(empresa_data['empresa'], 'telefone') and empresa_data['empresa'].telefone else 'N/A', data_format)
            empresas_sheet.write(row, 5, empresa_data['total_usuarios'], data_format)
            empresas_sheet.write(row, 6, empresa_data['total_chamados'], data_format)
            empresas_sheet.write(row, 7, empresa_data['chamados_abertos'], data_format)
            empresas_sheet.write(row, 8, empresa_data['chamados_andamento'], data_format)
            empresas_sheet.write(row, 9, empresa_data['chamados_finalizados'], data_format)
            
            taxa = (empresa_data['chamados_finalizados']/empresa_data['total_chamados']*100) if empresa_data['total_chamados'] > 0 else 0
            empresas_sheet.write(row, 10, f"{taxa:.1f}%", data_format)
            
            # Lista dos técnicos que atenderam
            tecnicos_names = []
            if 'tecnicos_atenderam' in empresa_data:
                tecnicos_names = [t['tecnico'].nome for t in empresa_data['tecnicos_atenderam']]
            empresas_sheet.write(row, 11, ', '.join(tecnicos_names) if tecnicos_names else 'Nenhum', data_format)
        
        # Aba de usuários por empresa
        usuarios_sheet = workbook.add_worksheet('Usuários por Empresa')
        
        # Cabeçalhos para usuários
        user_headers = ['Empresa', 'Usuário', 'Email', 'Telefone', 'Tipo', 'Total Chamados', 
                       'Abertos', 'Em Andamento', 'Finalizados', 'Taxa Resolução (%)']
        
        for i, header in enumerate(user_headers):
            usuarios_sheet.write(0, i, header, header_format)
        
        # Dados dos usuários
        user_row = 1
        for empresa_data in relatorio_data:
            if empresa_data['usuarios_stats']:
                for user_stat in empresa_data['usuarios_stats']:
                    if user_stat['total_chamados'] > 0:  # Só mostrar usuários com chamados
                        usuarios_sheet.write(user_row, 0, empresa_data['empresa'].nome_empresa, data_format)
                        usuarios_sheet.write(user_row, 1, user_stat['usuario'].nome, data_format)
                        usuarios_sheet.write(user_row, 2, user_stat['usuario'].email, data_format)
                        usuarios_sheet.write(user_row, 3, user_stat['usuario'].telefone if hasattr(user_stat['usuario'], 'telefone') and user_stat['usuario'].telefone else 'N/A', data_format)
                        usuarios_sheet.write(user_row, 4, user_stat['usuario'].tipo_usuario if hasattr(user_stat['usuario'], 'tipo_usuario') else 'N/A', data_format)
                        usuarios_sheet.write(user_row, 5, user_stat['total_chamados'], data_format)
                        usuarios_sheet.write(user_row, 6, user_stat['abertos'], data_format)
                        usuarios_sheet.write(user_row, 7, user_stat['em_andamento'], data_format)
                        usuarios_sheet.write(user_row, 8, user_stat['finalizados'], data_format)
                        
                        taxa_user = (user_stat['finalizados']/user_stat['total_chamados']*100) if user_stat['total_chamados'] > 0 else 0
                        usuarios_sheet.write(user_row, 9, f"{taxa_user:.1f}%", data_format)
                        user_row += 1
        
        # Ajustar largura das colunas
        empresas_sheet.set_column('A:A', 25)
        empresas_sheet.set_column('B:B', 15)
        empresas_sheet.set_column('C:C', 20)
        empresas_sheet.set_column('D:E', 12)
        empresas_sheet.set_column('F:L', 12)
        
        usuarios_sheet.set_column('A:B', 25)
        usuarios_sheet.set_column('C:C', 30)
        usuarios_sheet.set_column('D:J', 12)
        
        # Aba de chamados detalhados
        chamados_sheet = workbook.add_worksheet('Chamados Detalhados')
        
        # Cabeçalhos para chamados
        chamados_headers = ['Empresa', 'Título', 'Status', 'Prioridade', 'Data Abertura', 
                           'Data Finalização', 'Usuário', 'Técnico', 'Tempo Resolução (dias)']
        
        for i, header in enumerate(chamados_headers):
            chamados_sheet.write(0, i, header, header_format)
        
        # Dados dos chamados
        chamado_row = 1
        for empresa_data in relatorio_data:
            if 'chamados_detalhados' in empresa_data and empresa_data['chamados_detalhados']:
                for chamado_detail in empresa_data['chamados_detalhados']:
                    chamados_sheet.write(chamado_row, 0, empresa_data['empresa'].nome_empresa, data_format)
                    chamados_sheet.write(chamado_row, 1, chamado_detail['titulo'], data_format)
                    chamados_sheet.write(chamado_row, 2, chamado_detail['status'].replace('_', ' ').title(), data_format)
                    chamados_sheet.write(chamado_row, 3, chamado_detail['prioridade'].title(), data_format)
                    
                    # Datas formatadas
                    data_abertura = chamado_detail['data_abertura'].strftime('%d/%m/%Y %H:%M') if chamado_detail['data_abertura'] else 'N/A'
                    data_finalizacao = chamado_detail['data_finalizacao'].strftime('%d/%m/%Y %H:%M') if chamado_detail['data_finalizacao'] else 'Em aberto'
                    
                    chamados_sheet.write(chamado_row, 4, data_abertura, data_format)
                    chamados_sheet.write(chamado_row, 5, data_finalizacao, data_format)
                    chamados_sheet.write(chamado_row, 6, chamado_detail['usuario'], data_format)
                    chamados_sheet.write(chamado_row, 7, chamado_detail['tecnico'], data_format)
                    
                    tempo_resolucao = chamado_detail['tempo_resolucao'] if chamado_detail['tempo_resolucao'] is not None else 'N/A'
                    chamados_sheet.write(chamado_row, 8, tempo_resolucao, data_format)
                    
                    chamado_row += 1
        
        # Ajustar largura das colunas para chamados
        chamados_sheet.set_column('A:A', 25)
        chamados_sheet.set_column('B:B', 40)
        chamados_sheet.set_column('C:D', 15)
        chamados_sheet.set_column('E:F', 18)
        chamados_sheet.set_column('G:H', 25)
        chamados_sheet.set_column('I:I', 15)
        
        workbook.close()
        buffer.seek(0)
        return buffer
    
    def export_tecnicos_pdf(self, relatorio_data):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Título
        title = Paragraph("Relatório de Técnicos", self.title_style)
        story.append(title)
        
        # Data de geração
        date_str = datetime.now().strftime('%d/%m/%Y às %H:%M')
        date_p = Paragraph(f"<i>Gerado em: {date_str}</i>", self.styles['Normal'])
        story.append(date_p)
        story.append(Spacer(1, 20))
        
        # Resumo geral
        total_tecnicos = len(relatorio_data)
        total_chamados = sum(item['total_chamados'] for item in relatorio_data)
        total_andamento = sum(item['chamados_andamento'] for item in relatorio_data)
        total_finalizados = sum(item['chamados_finalizados'] for item in relatorio_data)
        
        resumo_data = [
            ['Métrica', 'Valor'],
            ['Total de Técnicos Ativos', str(total_tecnicos)],
            ['Total de Chamados Atendidos', str(total_chamados)],
            ['Chamados em Andamento', str(total_andamento)],
            ['Chamados Finalizados', str(total_finalizados)],
            ['Taxa de Resolução Geral', f"{(total_finalizados/total_chamados*100) if total_chamados > 0 else 0:.1f}%"]
        ]
        
        resumo_table = Table(resumo_data)
        resumo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(Paragraph("Resumo Geral", self.subtitle_style))
        story.append(resumo_table)
        story.append(Spacer(1, 30))
        
        # Detalhes por técnico
        story.append(Paragraph("Detalhes por Técnico", self.subtitle_style))
        
        for tecnico_data in relatorio_data:
            # Nome do técnico
            tecnico_title = Paragraph(f"<b>{tecnico_data['tecnico'].nome}</b>", 
                                    ParagraphStyle('TecnicoTitle', fontSize=12, textColor=colors.darkgreen))
            story.append(tecnico_title)
            
            # Info básica do técnico
            info_lines = [
                f"Email: {tecnico_data['tecnico'].email}",
                f"Telefone: {tecnico_data['tecnico'].telefone if hasattr(tecnico_data['tecnico'], 'telefone') and tecnico_data['tecnico'].telefone else 'N/A'}",
                f"Data de Cadastro: {tecnico_data['tecnico'].data_criacao.strftime('%d/%m/%Y') if hasattr(tecnico_data['tecnico'], 'data_criacao') else 'N/A'}",
                f"Tipo de Usuário: {tecnico_data['tecnico'].tipo_usuario if hasattr(tecnico_data['tecnico'], 'tipo_usuario') else 'N/A'}"
            ]
            for info_line in info_lines:
                story.append(Paragraph(info_line, self.styles['Normal']))
            story.append(Spacer(1, 10))
            
            # Estatísticas do técnico
            stats_data = [
                ['Métrica', 'Quantidade'],
                ['Total Atendidos', str(tecnico_data['total_chamados'])],
                ['Chamados Abertos', str(tecnico_data['chamados_abertos'])],
                ['Chamados em Andamento', str(tecnico_data['chamados_andamento'])],
                ['Chamados Finalizados', str(tecnico_data['chamados_finalizados'])],
                ['Taxa de Resolução', f"{(tecnico_data['chamados_finalizados']/tecnico_data['total_chamados']*100) if tecnico_data['total_chamados'] > 0 else 0:.1f}%"]
            ]
            
            stats_table = Table(stats_data)
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(stats_table)
            
            # Lista detalhada de chamados do técnico (últimos 10)
            if 'chamados_detalhados' in tecnico_data and tecnico_data['chamados_detalhados']:
                story.append(Spacer(1, 15))
                story.append(Paragraph("<b>Últimos 10 Chamados Atendidos:</b>", self.styles['Normal']))
                
                chamados_data = [['Título', 'Status', 'Prioridade', 'Data Abertura', 'Data Finalização', 'Usuário', 'Tempo (dias)']]
                # Mostrar apenas os últimos 10 chamados, ordenados por data
                ultimos_chamados = sorted(tecnico_data['chamados_detalhados'], key=lambda x: x['data_abertura'], reverse=True)[:10]
                
                for chamado_detail in ultimos_chamados:
                    data_abertura = chamado_detail['data_abertura'].strftime('%d/%m/%Y %H:%M') if chamado_detail['data_abertura'] else 'N/A'
                    data_finalizacao = chamado_detail['data_finalizacao'].strftime('%d/%m/%Y %H:%M') if chamado_detail['data_finalizacao'] else 'Em aberto'
                    tempo_resolucao = f"{chamado_detail['tempo_resolucao']}" if chamado_detail['tempo_resolucao'] is not None else 'N/A'
                    
                    chamados_data.append([
                        chamado_detail['titulo'][:35] + '...' if len(chamado_detail['titulo']) > 35 else chamado_detail['titulo'],
                        chamado_detail['status'].replace('_', ' ').title(),
                        chamado_detail['prioridade'].title(),
                        data_abertura,
                        data_finalizacao,
                        chamado_detail['usuario'][:25] if len(chamado_detail['usuario']) > 25 else chamado_detail['usuario'],
                        tempo_resolucao
                    ])
                
                if len(chamados_data) > 1:
                    chamados_table = Table(chamados_data)
                    chamados_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 7),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.lightcyan),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(chamados_table)
            
            story.append(Spacer(1, 30))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def export_tecnicos_excel(self, relatorio_data):
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        
        # Formatos
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#2e7d2e',
            'font_color': 'white'
        })
        
        header_format = workbook.add_format({
            'bold': True,
            'font_size': 12,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#5cb85c',
            'font_color': 'white',
            'border': 1
        })
        
        data_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        # Aba Resumo
        resumo_sheet = workbook.add_worksheet('Resumo Geral')
        
        resumo_sheet.write('A1', 'RELATÓRIO DE TÉCNICOS', title_format)
        resumo_sheet.merge_range('A1:F1', 'RELATÓRIO DE TÉCNICOS', title_format)
        
        # Data de geração
        date_str = datetime.now().strftime('%d/%m/%Y às %H:%M')
        resumo_sheet.write('A2', f'Gerado em: {date_str}')
        
        # Estatísticas gerais
        resumo_sheet.write('A4', 'Métrica', header_format)
        resumo_sheet.write('B4', 'Valor', header_format)
        
        total_tecnicos = len(relatorio_data)
        total_chamados = sum(item['total_chamados'] for item in relatorio_data)
        total_andamento = sum(item['chamados_andamento'] for item in relatorio_data)
        total_finalizados = sum(item['chamados_finalizados'] for item in relatorio_data)
        
        metrics = [
            ['Total de Técnicos Ativos', total_tecnicos],
            ['Total de Chamados Atendidos', total_chamados],
            ['Chamados em Andamento', total_andamento],
            ['Chamados Finalizados', total_finalizados],
            ['Taxa de Resolução Geral', f"{(total_finalizados/total_chamados*100) if total_chamados > 0 else 0:.1f}%"]
        ]
        
        for i, (metric, value) in enumerate(metrics, 5):
            resumo_sheet.write(f'A{i}', metric, data_format)
            resumo_sheet.write(f'B{i}', value, data_format)
        
        # Aba por técnico
        tecnicos_sheet = workbook.add_worksheet('Técnicos Detalhado')
        
        # Cabeçalhos
        headers = ['Técnico', 'Email', 'Telefone', 'Data Cadastro', 'Tipo Usuário', 'Total Atendidos', 
                  'Abertos', 'Em Andamento', 'Finalizados', 'Taxa Resolução (%)', 'Média Diária', 'Tempo Médio Resolução']
        
        for i, header in enumerate(headers):
            tecnicos_sheet.write(0, i, header, header_format)
        
        # Dados dos técnicos
        for row, tecnico_data in enumerate(relatorio_data, 1):
            tecnicos_sheet.write(row, 0, tecnico_data['tecnico'].nome, data_format)
            tecnicos_sheet.write(row, 1, tecnico_data['tecnico'].email, data_format)
            tecnicos_sheet.write(row, 2, tecnico_data['tecnico'].telefone if hasattr(tecnico_data['tecnico'], 'telefone') and tecnico_data['tecnico'].telefone else 'N/A', data_format)
            tecnicos_sheet.write(row, 3, tecnico_data['tecnico'].data_criacao.strftime('%d/%m/%Y') if hasattr(tecnico_data['tecnico'], 'data_criacao') else 'N/A', data_format)
            tecnicos_sheet.write(row, 4, tecnico_data['tecnico'].tipo_usuario if hasattr(tecnico_data['tecnico'], 'tipo_usuario') else 'N/A', data_format)
            tecnicos_sheet.write(row, 5, tecnico_data['total_chamados'], data_format)
            tecnicos_sheet.write(row, 6, tecnico_data['chamados_abertos'], data_format)
            tecnicos_sheet.write(row, 7, tecnico_data['chamados_andamento'], data_format)
            tecnicos_sheet.write(row, 8, tecnico_data['chamados_finalizados'], data_format)
            
            taxa = (tecnico_data['chamados_finalizados']/tecnico_data['total_chamados']*100) if tecnico_data['total_chamados'] > 0 else 0
            tecnicos_sheet.write(row, 9, f"{taxa:.1f}%", data_format)
            
            # Calcular média diária (assumindo dados dos últimos 30 dias)
            media_diaria = tecnico_data['total_chamados'] / 30.0
            tecnicos_sheet.write(row, 10, f"{media_diaria:.1f}", data_format)
            
            # Tempo médio de resolução (placeholder - pode ser calculado se tiver dados de tempo)
            tecnicos_sheet.write(row, 11, "N/A", data_format)
        
        # Aba de desempenho detalhado
        performance_sheet = workbook.add_worksheet('Desempenho Técnicos')
        
        # Cabeçalhos para desempenho
        perf_headers = ['Técnico', 'Total Chamados', 'Finalizados', 'Taxa Sucesso (%)', 
                       'Produtividade', 'Classificação', 'Observações']
        
        for i, header in enumerate(perf_headers):
            performance_sheet.write(0, i, header, header_format)
        
        # Dados de desempenho
        sorted_tecnicos = sorted(relatorio_data, key=lambda x: x['chamados_finalizados'], reverse=True)
        for row, tecnico_data in enumerate(sorted_tecnicos, 1):
            performance_sheet.write(row, 0, tecnico_data['tecnico'].nome, data_format)
            performance_sheet.write(row, 1, tecnico_data['total_chamados'], data_format)
            performance_sheet.write(row, 2, tecnico_data['chamados_finalizados'], data_format)
            
            taxa = (tecnico_data['chamados_finalizados']/tecnico_data['total_chamados']*100) if tecnico_data['total_chamados'] > 0 else 0
            performance_sheet.write(row, 3, f"{taxa:.1f}%", data_format)
            
            # Calcular produtividade (baseada no número de chamados finalizados)
            produtividade = "Alta" if tecnico_data['chamados_finalizados'] > 10 else "Média" if tecnico_data['chamados_finalizados'] > 5 else "Baixa"
            performance_sheet.write(row, 4, produtividade, data_format)
            
            # Classificação baseada na posição
            classificacao = f"{row}º lugar"
            performance_sheet.write(row, 5, classificacao, data_format)
            
            # Observações
            obs = []
            if taxa >= 90:
                obs.append("Excelente taxa de resolução")
            elif taxa >= 70:
                obs.append("Boa taxa de resolução")
            else:
                obs.append("Pode melhorar a taxa de resolução")
                
            if tecnico_data['chamados_andamento'] > 5:
                obs.append("Muitos chamados em andamento")
                
            performance_sheet.write(row, 6, "; ".join(obs) if obs else "Desempenho regular", data_format)
        
        # Ajustar largura das colunas
        tecnicos_sheet.set_column('A:A', 25)
        tecnicos_sheet.set_column('B:B', 25)
        tecnicos_sheet.set_column('C:L', 15)
        
        performance_sheet.set_column('A:A', 25)
        performance_sheet.set_column('B:F', 15)
        performance_sheet.set_column('G:G', 40)
        
        # Aba de chamados atendidos por técnico
        chamados_tecnicos_sheet = workbook.add_worksheet('Chamados por Técnico')
        
        # Cabeçalhos para chamados por técnico
        chamados_headers = ['Técnico', 'Título do Chamado', 'Status', 'Prioridade', 'Data Abertura', 
                           'Data Finalização', 'Cliente', 'Tempo Resolução (dias)']
        
        for i, header in enumerate(chamados_headers):
            chamados_tecnicos_sheet.write(0, i, header, header_format)
        
        # Dados dos chamados por técnico
        chamado_row = 1
        for tecnico_data in relatorio_data:
            if 'chamados_detalhados' in tecnico_data and tecnico_data['chamados_detalhados']:
                for chamado_detail in tecnico_data['chamados_detalhados']:
                    chamados_tecnicos_sheet.write(chamado_row, 0, tecnico_data['tecnico'].nome, data_format)
                    chamados_tecnicos_sheet.write(chamado_row, 1, chamado_detail['titulo'], data_format)
                    chamados_tecnicos_sheet.write(chamado_row, 2, chamado_detail['status'].replace('_', ' ').title(), data_format)
                    chamados_tecnicos_sheet.write(chamado_row, 3, chamado_detail['prioridade'].title(), data_format)
                    
                    # Datas formatadas
                    data_abertura = chamado_detail['data_abertura'].strftime('%d/%m/%Y %H:%M') if chamado_detail['data_abertura'] else 'N/A'
                    data_finalizacao = chamado_detail['data_finalizacao'].strftime('%d/%m/%Y %H:%M') if chamado_detail['data_finalizacao'] else 'Em aberto'
                    
                    chamados_tecnicos_sheet.write(chamado_row, 4, data_abertura, data_format)
                    chamados_tecnicos_sheet.write(chamado_row, 5, data_finalizacao, data_format)
                    chamados_tecnicos_sheet.write(chamado_row, 6, chamado_detail['usuario'], data_format)
                    
                    tempo_resolucao = chamado_detail['tempo_resolucao'] if chamado_detail['tempo_resolucao'] is not None else 'N/A'
                    chamados_tecnicos_sheet.write(chamado_row, 7, tempo_resolucao, data_format)
                    
                    chamado_row += 1
        
        # Ajustar largura das colunas para chamados
        chamados_tecnicos_sheet.set_column('A:A', 25)
        chamados_tecnicos_sheet.set_column('B:B', 40)
        chamados_tecnicos_sheet.set_column('C:D', 15)
        chamados_tecnicos_sheet.set_column('E:F', 18)
        chamados_tecnicos_sheet.set_column('G:G', 25)
        chamados_tecnicos_sheet.set_column('H:H', 15)
        
        workbook.close()
        buffer.seek(0)
        return buffer